"""Final project for Group 4 (Peter Koppelman, Michael Shore)
 for Marc Bacchus' programming with python course at NYU SPS
December 17, 2018."""

import auth_token as at
import email_the_data
import openpyxl
import numpy as np
import pandas as pd
import quandl

from openpyxl.chart import Reference, LineChart
from openpyxl import load_workbook
from openpyxl.chart.axis import DateAxis
from pandas import ExcelWriter
from pandas.tseries.offsets import BQuarterEnd
from win32com.client import Dispatch
from pathlib import Path
from datetime import timedelta
from dateutil.easter import *


def treasury_FM_data_pull():

    def get_data():
        # Get treasury yield data from Quandl
        df_treasury = quandl.get("USTREASURY/YIELD", authtoken=at.auth_key)

        # This time we get daily change in interest rate data
        df_deltas = quandl.get("USTREASURY/YIELD", authtoken=at.auth_key,
                               transform="diff").fillna(0)

        # Get Freddie, conforming and Jumbo 30 yr mortgage data from Quandl
        # Get Freddie Mac Data
        df_freddie_mac = quandl.get("FMAC/30US", authtoken=at.auth_key)

        # Get Wells Fargo conforming data
        df_wfc_conf_30 = quandl.get("WFC/PR_CON_30YFIXED_IR",
                                 authtoken=at.auth_key)

        # Get Wells Fargo Jumbo data
        df_wfc_jumbo_30 = quandl.get("WFC/PR_JUMBO_30YFIXED_IR",
                                  authtoken=at.auth_key)

        return df_treasury, df_deltas, df_freddie_mac, df_wfc_conf_30, df_wfc_jumbo_30

    def manipulate_mortgage_data(df_freddie_mac, df_wfc_conf_30, df_wfc_jumbo_30):
        # Rename columns
        df_freddie_mac.columns = ['Freddie Mac']
        df_wfc_conf_30.columns = ['WFC Conf Int Rate']
        df_wfc_jumbo_30.columns = ['WFC Jumbo Int Rate']

        # Merge freddie mac, Wells Fargo conforming and Wells Fargo jumbo datafrmes
        df_mortgage = pd.merge(df_freddie_mac, df_wfc_jumbo_30, 
            left_index = True, right_index = True)
        df_mortgage = pd.merge(df_mortgage, df_wfc_conf_30, 
            left_index = True, right_index = True)
        
        # Get differences in rates. Add them to the dataframe
        df_mortgage['Conf minus Freddie'] = df_mortgage['WFC Conf Int Rate'] - \
        df_mortgage['Freddie Mac']
        df_mortgage['Jumbo minus Freddie'] = df_mortgage['WFC Jumbo Int Rate'] - \
        df_mortgage['Freddie Mac']
        df_mortgage.reset_index(inplace=True)
        return df_mortgage, len(df_mortgage.index)


    def manipulate_deltas_data(df_deltas):
        # Drop the 2 month index. 2 month data only started appearing recently
        # Takes the last 5 years worth of data
        df_deltas = df_deltas.drop(['2 MO'], axis=1).last('5Y')
        df_deltas.reset_index(inplace=True)
        return df_deltas, len(df_deltas.index)

    def manipulate_yield_data(df_treasury):
        # Take out 2 month yield and only use last two years data
        df_treasury = df_treasury.drop(['2 MO'], axis=1).last('2Y')
        df_treasury.reset_index(inplace=True)

        # Get data from last day of each business quarter
        # Add data from the most recent day we have
        # Transpose the dataframe
        df_us_treasury = df_treasury.loc[df_treasury.Date.isin(df_treasury.Date + BQuarterEnd())]

        # Once every 5 or 6 years the end of the first business
        # quarter is Good Friday. The financial markets in the US 
        # are closed on Good Friday. When this occurs we have to get
        # data from the day before Good Friday.

        # Create a list of the years that are in the df_treasury dataframe 
        df_dates = pd.to_datetime(df_treasury['Date']).dt.year.unique()

        for date in df_dates:
            # Calculate Good Friday. It's two days before Easter Sunday
            goodfriday = easter(date) + timedelta(days = -2)
            # Calculate the end of the business quarter for the quarter that 
            # Good Friday is in.
            Bqtr_end_date = (pd.to_datetime(goodfriday) + BQuarterEnd(0)).date()

            # check to see if Good Friday is the last day of the business quarter
            if goodfriday == Bqtr_end_date:

                # Subtract one day from Good Friday to get financial end of qtr
                end_of_qtr = pd.to_datetime(goodfriday + timedelta(days = -1))
                # Get the row in df_treasury with the information that we need
                df_temp = df_treasury[df_treasury.Date == end_of_qtr]
                # Add the dataframe with the one record that we need to the 
                # dataframe with the business quarter end data
                df_us_treasury = pd.concat([df_us_treasury, df_temp])


        df_us_treasury = (df_us_treasury.append(df_treasury.iloc[-1], ignore_index=True)).T
        return df_us_treasury, len(df_us_treasury.index), len(df_us_treasury.columns) + 1


    def export_to_excel(df_mortgage, df_deltas, df_us_treasury):
        # Export to Excel for use with openpyxl
        writer = ExcelWriter(at.file_name, engine='openpyxl')
        df_mortgage.to_excel(writer, index=False, sheet_name='mortgage_rates')
        df_deltas.to_excel(writer, index=False, sheet_name=
                           'treasury_delta_data')
        df_us_treasury.to_excel(writer, index=True,
                             sheet_name='us_treasury_data')
        writer.save()
        writer.close()


    def create_charts(max_row_mortgage, max_row_delta, max_row_treasury, max_col_treasury):
        def mortgage_graphs(max_row_mortgage, wb):

            ws = wb['mortgage_rates']
            ws1 = wb.create_sheet('mortgage_graph', index = 0)
            """Line Chart."""
            data = Reference(ws, min_col=2, min_row=1, max_col=4,
                             max_row = max_row_mortgage)
            dates = Reference(ws, min_col = 1, min_row = 2,
                              max_row = max_row_mortgage)

            c1 = LineChart()
            c1.title = "Freddie Mac vs Conforming and Non-Conforming Interest Rates"
            c1.y_axis.title = "Interest Rate"
            c1.y_axis.crossAx = 500
            c1.x_axis = DateAxis(crossAx=100)
            c1.x_axis.number_format = 'mm-dd-yyyy'
            c1.x_axis.majorTimeUnit = "days"
            c1.x_axis.title = "Date"
            c1.height = 15
            c1.width = 30

            c1.add_data(data, titles_from_data=True)
            c1.set_categories(dates)
            ws1.add_chart(c1, "A1")

            """Deltas for line chart."""
            ws1 = wb.create_sheet('mortgage_delta_graph', index = 1)
            data = Reference(ws, min_col=5, min_row=1,
                             max_col=6, max_row=max_row_mortgage)
            c2 = LineChart()
            c2.title = "Difference in Basis Points"
            c2.y_axis.title = "Basis Points"
            c2.y_axis.crossAx = 500
            c2.x_axis = DateAxis(crossAx=100)
            c2.x_axis.number_format = 'mm-dd-yyyy'
            c2.x_axis.majorTimeUnit = "days"
            c2.x_axis.title = "Date"
            c2.height = 15
            c2.width = 30
            c2.x_axis.tickLblPos = "low"

            c2.add_data(data, titles_from_data=True)
            c2.set_categories(dates)
            ws1.add_chart(c2, "A1")

            # Saves workbook.
            wb.save(at.file_name)

        def treasury_delta_data(max_row_delta, wb):
            """Treasury Delta Data."""
            # Make sure that there are the same number of items in the list of
            # titles as well as the list of chart cells (top left hand cell 
            # for each graph)
            if len(at.title) != len(at.chart_cell):
                print('The length of the title list is not same as \
                          the length of the chart_cell list.')
                return

            ws = wb['treasury_delta_data']
            ws1 = wb.create_sheet('delta_graphs', index = 2)
            dates = Reference(ws, min_col=1, min_row=2, max_row=max_row_delta)

            for i in range(len(at.title)): 
                data = Reference(ws, min_col=i + 2, min_row=1, max_row=max_row_delta)
                c1 = LineChart()
                c1.x_axis.title = "Date"
                c1.y_axis.title = "Daily Delta"
                c1.y_axis.crossAx = 500
                c1.x_axis = DateAxis(crossAx=100)
                c1.x_axis.number_format = 'mm-dd-yyyy'
                c1.x_axis.majorTimeUnit = "days"
                c1.x_axis.tickLblPos = "low"

                c1.add_data(data, titles_from_data=True)
                c1.set_categories(dates)
                c1.legend = None

                # Get title information and top left hand corner cell info
                # from the lists in at.py
                c1.title = at.title[i]
                ws1.add_chart(c1, at.chart_cell[i])

            # Save workbook
            wb.save(at.file_name)

        def treasury_yield_graphs(max_row_treasury, max_col_treasury, wb):
            """US Treasury Data for graph of yield curves."""
            ws = wb['us_treasury_data']
            ws1 = wb.create_sheet('treasury_graph', index = 3)

            # Fix date format in second column
            for col in range(2, max_col_treasury + 1):
                cell = ws.cell(row = 2, column = col)
                cell.value = cell.value.strftime('%Y-%m-%d')

            c1 = LineChart()
            c1.title = "Treasury Yields"
            c1.x_axis.title = "Tenor"
            c1.y_axis.title = "Interest Rate"

            data = Reference(ws, min_col=2, min_row=2, max_col= max_col_treasury,
                             max_row=max_row_treasury+1)
            c1.add_data(data, titles_from_data=True)
            yvalues = Reference(ws, min_col=1, min_row=3,
                                max_row = max_row_treasury+1)
            c1.set_categories(yvalues)
            c1.height = 15
            c1.width = 30
            ws1.add_chart(c1, "A1")

            # Save workbook
            wb.save(at.file_name)

 
        """Call functions in the create charts function."""
        # Open the workbook
        wb = load_workbook(filename = at.file_name)
        mortgage_graphs(max_row_mortgage, wb)
        treasury_delta_data(max_row_delta, wb)
        treasury_yield_graphs(max_row_treasury, max_col_treasury, wb)

        # Close the workbook
        wb.close()


    ## Run the Code. ##
    # Get the data from Quandl
    df_treasury, df_deltas, df_freddie_mac, df_wfc_conf_30, df_wfc_jumbo_30 = get_data()
    # Manipulate the data for use with openpyxl
    df_mortgage, max_row_mortgage = manipulate_mortgage_data(df_freddie_mac,
                                                             df_wfc_conf_30, 
                                                             df_wfc_jumbo_30)
    df_us_treasury, max_row_treasury, max_col_treasury  = manipulate_yield_data(df_treasury)
    df_deltas, max_row_delta, = manipulate_deltas_data(df_deltas)
    # Export to Excel using pandas library Excelwriter
    export_to_excel(df_mortgage, df_deltas, df_us_treasury)
    # Create the charts
    create_charts(max_row_mortgage, max_row_delta, max_row_treasury, max_col_treasury)
    # Email the data
    email_the_data.email_my_data()


if __name__ == '__main__':
    treasury_FM_data_pull()
